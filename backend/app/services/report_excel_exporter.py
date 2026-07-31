"""
Report Excel Export Service

Tạo file Excel (.xlsx) cho các báo cáo:
- Hợp đồng
- Hợp đồng sắp hết hạn
- Doanh thu tổng hợp

Note: Phase này chưa có revenue breakdown theo lĩnh vực.
Sheet "Theo_linh_vuc" chỉ thống kê số lượng HĐ.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# =============================================================================
# Date Formatting Helper
# =============================================================================

def parse_date_value(value: Union[str, date, datetime, None]) -> Optional[date]:
    """
    Parse various date formats to Python date object.
    
    Args:
        value: Can be date, datetime, ISO string like "2026-05-15", "2026-05-15T00:00:00", or None
    
    Returns:
        Python date object or None if cannot parse
    """
    if value is None:
        return None
    
    if isinstance(value, datetime):
        return value.date()
    
    if isinstance(value, date):
        return value
    
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        
        # Handle ISO format with time: "2026-05-15T00:00:00"
        if "T" in value:
            value = value.split("T")[0]
        
        # Try parsing various formats
        formats = [
            "%Y-%m-%d",      # 2026-05-15
            "%d/%m/%Y",      # 15/05/2026
            "%d-%m-%Y",      # 15-05-2026
            "%Y/%m/%d",      # 2026/05/15
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        
        logger.warning(f"Could not parse date string: '{value}'")
        return None
    
    return None


def format_date_cell(cell, value: Union[str, date, datetime, None], fallback_text: str = "-"):
    """
    Format a cell with proper date value and dd/mm/yyyy number format.
    
    Args:
        cell: openpyxl Cell object
        value: Date value in various formats or None
        fallback_text: Text to display if value is None/empty
    """
    parsed_date = parse_date_value(value)
    
    if parsed_date is None:
        cell.value = fallback_text
    else:
        cell.value = parsed_date
        cell.number_format = "dd/mm/yyyy"


# =============================================================================
# Domain Label Normalization
# =============================================================================

# Mapping raw domain labels to standardized display names
DOMAIN_NORMALIZE_MAP: dict[str, str] = {
    # Karaoke variants
    "karaoke": "Karaoke",
    "KARAOKE": "Karaoke",
    "Karoake": "Karaoke",
    
    # Phòng Thu Âm variants
    "phòng thu âm": "Phòng Thu Âm",
    "phòng thu am": "Phòng Thu Âm",
    "Phòng Thu Âm": "Phòng Thu Âm",
    "Phòng thu âm": "Phòng Thu Âm",
    "Phòng Thu Am": "Phòng Thu Âm",
    "PTA": "Phòng Thu Âm",
    "pta": "Phòng Thu Âm",
    "phòng ghi âm": "Phòng Thu Âm",
    "Phòng ghi âm": "Phòng Thu Âm",
    "phong thu am": "Phòng Thu Âm",
    
    # Nhà Hàng variants
    "nhà hàng": "Nhà Hàng",
    "nha hang": "Nhà Hàng",
    "Nhà Hàng": "Nhà Hàng",
    "Nha Hang": "Nhà Hàng",
    "NHÀ HÀNG": "Nhà Hàng",
    
    # Khu Vui Chơi variants
    "khu vui chơi": "Khu Vui Chơi",
    "khu vui choi": "Khu Vui Chơi",
    "Khu Vui Chơi": "Khu Vui Chơi",
    "Khu Vui Choi": "Khu Vui Chơi",
    "KVC": "Khu Vui Chơi",
    "kvc": "Khu Vui Chơi",
    
    # Chăm Sóc Sức Khoẻ variants
    "chăm sóc sức khoẻ": "Chăm Sóc Sức Khoẻ",
    "cham soc suc khoe": "Chăm Sóc Sức Khoẻ",
    "Chăm Sóc Sức Khoẻ": "Chăm Sóc Sức Khoẻ",
    "CSSK": "Chăm Sóc Sức Khoẻ",
    "cssk": "Chăm Sóc Sức Khoẻ",
    
    # Cà Phê variants
    "cà phê": "Cà Phê",
    "ca phe": "Cà Phê",
    "Cà Phê": "Cà Phê",
    "Ca Phe": "Cà Phê",
    "Cà phê": "Cà Phê",
    "cafe": "Cà Phê",
    "CAFE": "Cà Phê",
    "Coffee": "Cà Phê",
    
    # Biểu Diễn variants
    "BD": "Biểu Diễn",
    "bd": "Biểu Diễn",
    "Biểu diễn": "Biểu Diễn",
    "biểu diễn": "Biểu Diễn",
    
    # Sân Khấu Trực Tuyến
    "SCTT": "Sân Khấu Trực Tuyến",
    "sctt": "Sân Khấu Trực Tuyến",
    "Sân Khấu Trực Tuyến": "Sân Khấu Trực Tuyến",
    "sân khấu trực tuyến": "Sân Khấu Trực Tuyến",
}


def normalize_domain_label_for_report(value: str) -> str:
    """
    Normalize domain/linh vuc label for consistent reporting.
    
    - Maps various raw labels to standardized display names
    - Returns original if no mapping found
    - Logs unmapped labels for review
    """
    if not value:
        return "Không xác định"
    
    normalized = DOMAIN_NORMALIZE_MAP.get(value.strip())
    if normalized:
        return normalized
    
    # Try case-insensitive match
    value_lower = value.strip().lower()
    for key, display in DOMAIN_NORMALIZE_MAP.items():
        if key.lower() == value_lower:
            return display
    
    # No mapping found - return original but log for review
    logger.warning(f"Unmapped domain label in report: '{value}'")
    return value.strip()


# =============================================================================
# Styles
# =============================================================================

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

ALT_FILL = PatternFill(start_color="F5F9FC", end_color="F5F9FC", fill_type="solid")

WARNING_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
WARNING_FONT = Font(name="Arial", size=10, bold=True, color="856404")

TOTAL_FONT = Font(name="Arial", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="E8F4E8", end_color="E8F4E8", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header(ws, row_num: int, headers: list[str], col_widths: list[int]):
    """Apply header row with styling."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_data_row(ws, row_num: int, values: list, is_alt: bool = False, is_warning: bool = False):
    """Apply data row with styling."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = WARNING_FONT if is_warning else DATA_FONT
        cell.fill = WARNING_FILL if is_warning else (ALT_FILL if is_alt else PatternFill())
        cell.border = BORDER_THIN
        
        # Auto-align based on content type
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.alignment = DATA_ALIGN_RIGHT
            if value > 0:
                cell.number_format = '#,##0'
        elif isinstance(value, str) and value.startswith("0"):
            cell.alignment = DATA_ALIGN_CENTER
        else:
            cell.alignment = DATA_ALIGN_LEFT


def _freeze_and_filter(ws, row: int = 2):
    """Freeze top row and enable auto-filter."""
    ws.freeze_panes = ws.cell(row=row, column=1)
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


# =============================================================================
# Contract Export
# =============================================================================

def export_contracts_report_xlsx(
    contracts: list[dict],
    filters: Optional[dict] = None,
    has_music_areas: bool = False,
    filename_suffix: str = "",
) -> BytesIO:
    """
    Export contracts list to Excel.

    Args:
        contracts: List of contract dicts with keys:
            - contract_no: str
            - customer_name: str
            - ten_bang_hieu: str
            - domain: str
            - status: str
            - start_date: str (ISO format)
            - end_date: str (ISO format)
            - so_tien_value: int (before VAT)
            - vat_rate: float
            - vat_amount: int
            - total: int (so_tien + VAT)
            - nguoi_thuc_hien: str
            - area_name: str (optional, for multi-area expansion)
            - scale_description: str (optional, summarized or expanded)
            - music_usage_type: str (optional)
            - note: str (optional)
            - gcn_no: str (optional, GCN certificate number)
        filters: Applied filters for reference
        has_music_areas: If True, includes music usage area columns.
        filename_suffix: Optional suffix appended to sheet title for mode differentiation.

    Returns:
        BytesIO buffer with Excel content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Hop_dong{filename_suffix}" if filename_suffix else "Hop_dong"

    # Base headers — includes Số GCN
    base_headers = [
        "STT", "Số HĐ", "Số GCN", "Năm", "Tên đơn vị", "Bảng hiệu", "Lĩnh vực",
        "Phường/Xã sử dụng", "Tỉnh/Thành sử dụng",
        "Trạng thái", "Ngày hiệu lực", "Ngày hết hạn",
        "Tiền trước thuế", "VAT (%)", "Tiền VAT", "Tổng tiền", "Người thực hiện"
    ]

    # Music usage area columns (added when has_music_areas=True)
    area_headers = [
        "Vị trí / khu vực sử dụng âm nhạc",
        "Số phòng / số chỗ",
        "Hình thức sử dụng âm nhạc",
        "Ghi chú khu vực",
    ]

    headers = base_headers + area_headers if has_music_areas else base_headers

    # Widen money columns to prevent ####### display
    base_col_widths = [5, 16, 18, 6, 28, 18, 16, 24, 20, 13, 13, 13, 22, 9, 18, 22, 20]
    area_col_widths = [30, 20, 25, 25]
    col_widths = base_col_widths + area_col_widths if has_music_areas else base_col_widths
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="BÁO CÁO DANH SÁCH HỢP ĐỒNG")
    title_cell.font = Font(name="Arial", bold=True, size=14, color="2D5F8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Filter info row
    if filters:
        filter_text = " | ".join([f"{k}: {v}" for k, v in filters.items() if v])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        filter_cell = ws.cell(row=2, column=1, value=f"Lọc: {filter_text}")
        filter_cell.font = Font(name="Arial", italic=True, size=9, color="666666")
        filter_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    header_row = 3 if filters else 2
    _apply_header(ws, header_row, headers, col_widths)
    ws.row_dimensions[header_row].height = 30
    
    # Data rows
    data_row = header_row + 1
    total_before_vat = 0
    total_vat = 0
    total_amount = 0
    base_col_count = 17  # Number of base columns (15 + usage_ward + usage_province)

    for idx, contract in enumerate(contracts, start=1):
        so_tien = contract.get("so_tien_value") or 0
        vat_rate = contract.get("vat_rate") or 0
        vat_amount = int(so_tien * vat_rate / 100) if vat_rate else 0
        total = so_tien + vat_amount

        total_before_vat += so_tien
        total_vat += vat_amount
        total_amount += total

        # Build row data
        start_date = contract.get("start_date", "")
        end_date = contract.get("end_date", "")
        status = contract.get("status", "-")
        row_data = [
            idx,
            contract.get("contract_no", "-"),
            contract.get("gcn_no", "-") or "-",
            contract.get("year", "-"),
            contract.get("customer_name", "-"),
            contract.get("ten_bang_hieu", "-"),
            contract.get("domain", "-"),
            contract.get("usage_ward") or "-",
            contract.get("usage_province") or "-",
            status.upper() if status else "-",
            start_date,  # col 11: Ngày hiệu lực
            end_date,    # col 12: Ngày hết hạn
            so_tien if so_tien > 0 else "-",
            f"{vat_rate:.1f}%" if vat_rate else "-",
            vat_amount if vat_amount > 0 else "-",
            total if total > 0 else "-",
            contract.get("nguoi_thuc_hien", "-"),
        ]

        # Add music usage area columns if enabled
        if has_music_areas:
            row_data.extend([
                contract.get("area_name", ""),
                contract.get("scale_description", ""),
                contract.get("music_usage_type", ""),
                contract.get("note", ""),
            ])

        # Write row with default styling
        _apply_data_row(ws, data_row, row_data, is_alt=(idx % 2 == 0))

        # Apply date formatting to Ngày hiệu lực (col 11) and Ngày hết hạn (col 12)
        format_date_cell(ws.cell(row=data_row, column=11), start_date)
        format_date_cell(ws.cell(row=data_row, column=12), end_date)
        data_row += 1

    # Totals row
    if contracts:
        # base_col_count=17 (15 base + usage_ward + usage_province)
        # without music_areas: merge_end=9 (cols 1-9), with music_areas: merge_end=13 (cols 1-13)
        merge_end = (15 + 4) if has_music_areas else 9
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=merge_end)
        total_label = ws.cell(row=data_row, column=1, value="TỔNG CỘNG")
        total_label.font = TOTAL_FONT
        total_label.fill = TOTAL_FILL
        total_label.alignment = DATA_ALIGN_RIGHT
        total_label.border = BORDER_THIN

        # Tiền trước thuế → merge_end+1
        ws.cell(row=data_row, column=merge_end + 1, value=total_before_vat).number_format = '#,##0'
        ws.cell(row=data_row, column=merge_end + 1).font = TOTAL_FONT
        ws.cell(row=data_row, column=merge_end + 1).fill = TOTAL_FILL
        ws.cell(row=data_row, column=merge_end + 1).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=merge_end + 1).border = BORDER_THIN

        # VAT %
        ws.cell(row=data_row, column=merge_end + 2, value="-")
        ws.cell(row=data_row, column=merge_end + 2).font = TOTAL_FONT
        ws.cell(row=data_row, column=merge_end + 2).fill = TOTAL_FILL
        ws.cell(row=data_row, column=merge_end + 2).alignment = DATA_ALIGN_CENTER
        ws.cell(row=data_row, column=merge_end + 2).border = BORDER_THIN

        # Tiền VAT → merge_end+3
        ws.cell(row=data_row, column=merge_end + 3, value=total_vat).number_format = '#,##0'
        ws.cell(row=data_row, column=merge_end + 3).font = TOTAL_FONT
        ws.cell(row=data_row, column=merge_end + 3).fill = TOTAL_FILL
        ws.cell(row=data_row, column=merge_end + 3).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=merge_end + 3).border = BORDER_THIN

        # Tổng tiền → merge_end+4
        ws.cell(row=data_row, column=merge_end + 4, value=total_amount).number_format = '#,##0'
        ws.cell(row=data_row, column=merge_end + 4).font = TOTAL_FONT
        ws.cell(row=data_row, column=merge_end + 4).fill = TOTAL_FILL
        ws.cell(row=data_row, column=merge_end + 4).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=merge_end + 4).border = BORDER_THIN

        # Fill remaining cells with "-"
        total_col_count = 15 + (4 if has_music_areas else 0)
        remaining_start = merge_end + 4
        for col_idx in range(remaining_start + 1, total_col_count + 1):
            ws.cell(row=data_row, column=col_idx, value="-")
            ws.cell(row=data_row, column=col_idx).font = TOTAL_FONT
            ws.cell(row=data_row, column=col_idx).fill = TOTAL_FILL
            ws.cell(row=data_row, column=col_idx).alignment = DATA_ALIGN_CENTER
            ws.cell(row=data_row, column=col_idx).border = BORDER_THIN
    
    # Freeze and auto-filter
    _freeze_and_filter(ws, row=header_row + 1)
    _auto_width(ws)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# Expiring Contracts Export
# =============================================================================

def export_expiring_contracts_xlsx(
    contracts: list[dict],
    filters: Optional[dict] = None,
) -> BytesIO:
    """
    Export expiring contracts to Excel.

    Args:
        contracts: List of contract dicts with keys:
            - id: int
            - contract_no: str
            - gcn_no: str
            - partner: str
            - field: str
            - expire_date: str (ISO format)
            - days_left: int
            - status: str
            - nguoi_phu_trach: str
            - so_tien_value: int (before VAT)
            - vat_rate: float
            - vat_amount: int
            - total: int
        filters: Applied filters for reference

    Returns:
        BytesIO buffer with Excel content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sap_het_han"

    # Headers: STT, Số HĐ, Số GCN, Tên đơn vị, Lĩnh vực, Ngày hết hạn, Còn lại, Trạng thái, Tiền trước thuế, VAT, Tổng tiền, Người phụ trách
    headers = [
        "STT", "Số HĐ", "Số GCN", "Tên đơn vị", "Lĩnh vực",
        "Ngày hết hạn", "Còn lại (ngày)", "Trạng thái",
        "Tiền trước thuế", "VAT", "Tổng tiền", "Người phụ trách"
    ]
    col_widths = [5, 16, 18, 28, 16, 13, 13, 14, 22, 9, 22, 20]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="BÁO CÁO HỢP ĐỒNG SẮP HẾT HẠN")
    title_cell.font = Font(name="Arial", bold=True, size=14, color="C0392B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Filter info row
    if filters:
        filter_text = " | ".join([f"{k}: {v}" for k, v in filters.items() if v])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        filter_cell = ws.cell(row=2, column=1, value=f"Lọc: {filter_text}")
        filter_cell.font = Font(name="Arial", italic=True, size=9, color="666666")
        filter_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = 3 if filters else 2
    _apply_header(ws, header_row, headers, col_widths)
    ws.row_dimensions[header_row].height = 30

    # Data rows
    data_row = header_row + 1
    total_before_vat = 0
    total_vat = 0
    total_amount = 0

    for idx, contract in enumerate(contracts, start=1):
        days_left = contract.get("days_left", 0)
        expire_date = contract.get("expire_date", "")
        so_tien = contract.get("so_tien_value") or 0
        vat_rate = contract.get("vat_rate") or 0
        vat_amount = contract.get("vat_amount") or 0
        total = contract.get("total") or 0

        total_before_vat += so_tien
        total_vat += vat_amount
        total_amount += total

        # Highlight if <= 30 days
        is_warning = days_left <= 30

        # Build row data
        row_data = [
            idx,
            contract.get("contract_no", "-"),
            contract.get("gcn_no", "-") or "-",
            contract.get("partner", "-"),
            contract.get("field", "-"),
            expire_date,
            days_left if days_left > 0 else "-",
            contract.get("status", "-").upper() if contract.get("status") else "-",
            so_tien if so_tien > 0 else "-",
            f"{vat_rate:.1f}%" if vat_rate else "-",
            vat_amount if vat_amount > 0 else "-",
            total if total > 0 else "-",
            contract.get("nguoi_phu_trach", "-"),
        ]

        _apply_data_row(ws, data_row, row_data, is_alt=(idx % 2 == 0), is_warning=is_warning)
        # Apply date formatting to Ngày hết hạn (col 6)
        format_date_cell(ws.cell(row=data_row, column=6), expire_date)
        data_row += 1

    # Total row
    if contracts:
        merge_end = 8
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=merge_end)
        total_label = ws.cell(row=data_row, column=1, value="TỔNG CỘNG")
        total_label.font = TOTAL_FONT
        total_label.fill = TOTAL_FILL
        total_label.alignment = Alignment(horizontal="right", vertical="center")
        total_label.border = BORDER_THIN

        # Tiền trước thuế → col 9
        ws.cell(row=data_row, column=9, value=total_before_vat).number_format = '#,##0'
        ws.cell(row=data_row, column=9).font = TOTAL_FONT
        ws.cell(row=data_row, column=9).fill = TOTAL_FILL
        ws.cell(row=data_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=data_row, column=9).border = BORDER_THIN

        # VAT % → col 10
        ws.cell(row=data_row, column=10, value="-")
        ws.cell(row=data_row, column=10).font = TOTAL_FONT
        ws.cell(row=data_row, column=10).fill = TOTAL_FILL
        ws.cell(row=data_row, column=10).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=data_row, column=10).border = BORDER_THIN

        # Tiền VAT → col 11
        ws.cell(row=data_row, column=11, value=total_vat).number_format = '#,##0'
        ws.cell(row=data_row, column=11).font = TOTAL_FONT
        ws.cell(row=data_row, column=11).fill = TOTAL_FILL
        ws.cell(row=data_row, column=11).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=data_row, column=11).border = BORDER_THIN

        # Tổng tiền → col 12
        ws.cell(row=data_row, column=12, value=total_amount).number_format = '#,##0'
        ws.cell(row=data_row, column=12).font = TOTAL_FONT
        ws.cell(row=data_row, column=12).fill = TOTAL_FILL
        ws.cell(row=data_row, column=12).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=data_row, column=12).border = BORDER_THIN

    # Freeze and auto-filter
    _freeze_and_filter(ws, row=header_row + 1)
    _auto_width(ws)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# Period Export Builder
# =============================================================================

def export_period_xlsx(
    summary_data: dict,
    contracts: list[dict],
    expiring_contracts: list[dict],
    filters: Optional[dict] = None,
) -> BytesIO:
    """
    Export period-based report to Excel with 4 sheets:
    - Sheet 1: TONG_HOP — summary KPIs
    - Sheet 2: HOP_DONG_TRONG_KY — contracts signed in period (one row per contract)
    - Sheet 3: THEO_LINH_VUC — breakdown by field
    - Sheet 4: SAP_HET_HAN — expiring contracts in the period

    Args:
        summary_data: dict with keys:
            period_label, generated_date, total_contracts, signed_in_period,
            active_count, expiring_count, expired_count, pending_renewal_count,
            total_before_vat, total_vat, total_after_vat, gcn_issued, gcn_draft
        contracts: list of contract dicts from _build_contract_dict
        expiring_contracts: list of expiring contract dicts from _build_expiring_dict
        filters: applied filters dict

    Returns:
        BytesIO buffer with Excel content
    """
    wb = Workbook()

    # ── Sheet 1: TONG_HOP ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "TONG_HOP"

    summary_headers = ["Chỉ tiêu", "Giá trị"]
    col_widths_summary = [35, 22]

    # Title
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = ws_summary.cell(row=1, column=1, value="BÁO CÁO THEO KỲ")
    title_cell.font = Font(name="Arial", bold=True, size=14, color="2D5F8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 25

    # Period row
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    period_label = summary_data.get("period_label", "")
    generated_date = summary_data.get("generated_date", "")
    period_cell = ws_summary.cell(
        row=2, column=1,
        value=f"Kỳ báo cáo: {period_label}  |  Ngày xuất: {generated_date}"
    )
    period_cell.font = Font(name="Arial", italic=True, size=10, color="555555")
    period_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[2].height = 20

    # Filter info
    if filters and any(k not in ("Kỳ báo cáo",) for k in filters):
        filter_parts = [f"{k}: {v}" for k, v in filters.items() if k != "Kỳ báo cáo"]
        if filter_parts:
            ws_summary.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
            filter_cell = ws_summary.cell(row=3, column=1, value=" | ".join(filter_parts))
            filter_cell.font = Font(name="Arial", italic=True, size=9, color="888888")
            filter_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_summary.row_dimensions[3].height = 18
            header_start = 4
        else:
            header_start = 3
    else:
        header_start = 3

    _apply_header(ws_summary, header_start, summary_headers, col_widths_summary)
    ws_summary.row_dimensions[header_start].height = 30

    # KPI rows
    kpi_rows = [
        ("Tổng hợp đồng", summary_data.get("total_contracts", 0)),
        ("HĐ ký mới trong kỳ", summary_data.get("signed_in_period", 0)),
        ("HĐ đang hiệu lực", summary_data.get("active_count", 0)),
        ("HĐ sắp hết hạn", summary_data.get("expiring_count", 0)),
        ("HĐ đã hết hạn", summary_data.get("expired_count", 0)),
        ("HĐ chờ gia hạn", summary_data.get("pending_renewal_count", 0)),
        ("Tổng doanh thu trước VAT", summary_data.get("total_before_vat", 0)),
        ("Tổng VAT", summary_data.get("total_vat", 0)),
        ("Tổng thanh toán", summary_data.get("total_after_vat", 0)),
        ("GCN đã cấp", summary_data.get("gcn_issued", 0)),
        ("GCN bản nháp / chưa cấp", summary_data.get("gcn_draft", 0)),
    ]

    row_idx = header_start + 1
    for label, value in kpi_rows:
        ws_summary.cell(row=row_idx, column=1, value=label).font = DATA_FONT
        ws_summary.cell(row=row_idx, column=1).border = BORDER_THIN
        ws_summary.cell(row=row_idx, column=1).alignment = DATA_ALIGN_LEFT

        is_money = label.startswith("Tổng")
        if isinstance(value, (int, float)) and value > 0:
            cell = ws_summary.cell(row=row_idx, column=2, value=value)
            if is_money:
                cell.number_format = '#,##0'
            cell.alignment = DATA_ALIGN_RIGHT
        else:
            ws_summary.cell(row=row_idx, column=2, value=value).alignment = DATA_ALIGN_RIGHT
        ws_summary.cell(row=row_idx, column=2).font = DATA_FONT
        ws_summary.cell(row=row_idx, column=2).border = BORDER_THIN
        row_idx += 1

    _auto_width(ws_summary)

    # ── Sheet 2: HOP_DONG_TRONG_KY ─────────────────────────────────────
    ws_contracts = wb.create_sheet(title="HOP_DONG_TRONG_KY")

    contract_headers = [
        "STT", "Số HĐ", "Số GCN", "Tên đơn vị", "Lĩnh vực",
        "Phường/Xã sử dụng", "Tỉnh/Thành sử dụng",
        "Ngày ký", "Ngày hiệu lực", "Ngày hết hạn",
        "Tiền trước VAT", "VAT", "Tổng tiền", "Nhân viên",
    ]
    contract_col_widths = [5, 16, 18, 28, 16, 24, 20, 13, 13, 13, 22, 9, 22, 22]

    ws_contracts.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(contract_headers))
    title_c = ws_contracts.cell(row=1, column=1, value=f"HỢP ĐỒNG TRONG KỲ — {period_label}")
    title_c.font = Font(name="Arial", bold=True, size=13, color="2D5F8A")
    title_c.alignment = Alignment(horizontal="center", vertical="center")
    ws_contracts.row_dimensions[1].height = 22

    _apply_header(ws_contracts, 2, contract_headers, contract_col_widths)
    ws_contracts.row_dimensions[2].height = 30

    data_row = 3
    total_before_vat = 0
    total_vat_sum = 0
    total_after_vat = 0

    for idx, contract in enumerate(contracts, start=1):
        so_tien = contract.get("so_tien_value") or 0
        vat_rate = contract.get("vat_rate") or 0
        vat_amt = contract.get("vat_amount") or 0
        total = contract.get("total") or 0
        total_before_vat += so_tien
        total_vat_sum += vat_amt
        total_after_vat += total

        row_data = [
            idx,
            contract.get("contract_no", "-") or "-",
            contract.get("gcn_no", "-") or "-",
            contract.get("customer_name", "-") or "-",
            contract.get("domain", "-") or "-",
            contract.get("usage_ward") or "-",
            contract.get("usage_province") or "-",
            contract.get("start_date", "-") or "-",
            contract.get("start_date", "-") or "-",
            contract.get("end_date", "-") or "-",
            so_tien if so_tien > 0 else "-",
            f"{vat_rate:.1f}%" if vat_rate else "-",
            total if total > 0 else "-",
            contract.get("nguoi_thuc_hien", "-") or "-",
        ]

        _apply_data_row(ws_contracts, data_row, row_data, is_alt=(idx % 2 == 0))
        # Date columns: 8 (Ngày ký), 9 (Ngày hiệu lực), 10 (Ngày hết hạn)
        for col in (8, 9, 10):
            format_date_cell(ws_contracts.cell(row=data_row, column=col), row_data[col - 1])
        data_row += 1

    # Total row
    if contracts:
        merge_end = 5
        ws_contracts.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=merge_end)
        total_label = ws_contracts.cell(row=data_row, column=1, value="TỔNG CỘNG")
        total_label.font = TOTAL_FONT
        total_label.fill = TOTAL_FILL
        total_label.alignment = DATA_ALIGN_RIGHT
        total_label.border = BORDER_THIN

        for col_idx in range(merge_end + 1, 6):
            ws_contracts.cell(row=data_row, column=col_idx, value="-").border = BORDER_THIN
            ws_contracts.cell(row=data_row, column=col_idx).fill = TOTAL_FILL

        # Ward → col 6
        ws_contracts.cell(row=data_row, column=6, value="-")
        ws_contracts.cell(row=data_row, column=6).font = TOTAL_FONT
        ws_contracts.cell(row=data_row, column=6).fill = TOTAL_FILL
        ws_contracts.cell(row=data_row, column=6).border = BORDER_THIN

        # Province → col 7
        ws_contracts.cell(row=data_row, column=7, value="-")
        ws_contracts.cell(row=data_row, column=7).font = TOTAL_FONT
        ws_contracts.cell(row=data_row, column=7).fill = TOTAL_FILL
        ws_contracts.cell(row=data_row, column=7).border = BORDER_THIN

        # Tiền trước VAT → col 11
        ws_contracts.cell(row=data_row, column=11, value=total_before_vat).number_format = '#,##0'
        ws_contracts.cell(row=data_row, column=11).font = TOTAL_FONT
        ws_contracts.cell(row=data_row, column=11).fill = TOTAL_FILL
        ws_contracts.cell(row=data_row, column=11).alignment = DATA_ALIGN_RIGHT
        ws_contracts.cell(row=data_row, column=11).border = BORDER_THIN

        # VAT % → col 12
        ws_contracts.cell(row=data_row, column=12, value="-")
        ws_contracts.cell(row=data_row, column=12).font = TOTAL_FONT
        ws_contracts.cell(row=data_row, column=12).fill = TOTAL_FILL
        ws_contracts.cell(row=data_row, column=12).alignment = DATA_ALIGN_CENTER
        ws_contracts.cell(row=data_row, column=12).border = BORDER_THIN

        # Tổng tiền → col 13
        ws_contracts.cell(row=data_row, column=13, value=total_after_vat).number_format = '#,##0'
        ws_contracts.cell(row=data_row, column=13).font = TOTAL_FONT
        ws_contracts.cell(row=data_row, column=13).fill = TOTAL_FILL
        ws_contracts.cell(row=data_row, column=13).alignment = DATA_ALIGN_RIGHT
        ws_contracts.cell(row=data_row, column=13).border = BORDER_THIN

        # Nhân viên → col 14
        ws_contracts.cell(row=data_row, column=14, value="-")
        ws_contracts.cell(row=data_row, column=14).font = TOTAL_FONT
        ws_contracts.cell(row=data_row, column=14).fill = TOTAL_FILL
        ws_contracts.cell(row=data_row, column=12).alignment = DATA_ALIGN_CENTER
        ws_contracts.cell(row=data_row, column=12).border = BORDER_THIN

    _freeze_and_filter(ws_contracts, row=3)
    _auto_width(ws_contracts)

    # ── Sheet 3: THEO_LINH_VUC ──────────────────────────────────────────
    ws_field = wb.create_sheet(title="THEO_LINH_VUC")

    field_headers = ["STT", "Lĩnh vực", "Số HĐ", "Doanh thu trước VAT", "VAT", "Tổng thanh toán"]
    field_col_widths = [6, 30, 10, 22, 12, 22]

    ws_field.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(field_headers))
    title_f = ws_field.cell(row=1, column=1, value=f"DOANH THU THEO LĨNH VỰC — {period_label}")
    title_f.font = Font(name="Arial", bold=True, size=13, color="2D5F8A")
    title_f.alignment = Alignment(horizontal="center", vertical="center")
    ws_field.row_dimensions[1].height = 22

    _apply_header(ws_field, 2, field_headers, field_col_widths)
    ws_field.row_dimensions[2].height = 30

    # Group by domain
    field_groups: dict[str, dict] = {}
    for c in contracts:
        dom = str(c.get("domain") or "Không xác định")
        if dom not in field_groups:
            field_groups[dom] = {"count": 0, "before_vat": 0, "vat": 0, "total": 0}
        field_groups[dom]["count"] += 1
        field_groups[dom]["before_vat"] += c.get("so_tien_value") or 0
        field_groups[dom]["vat"] += c.get("vat_amount") or 0
        field_groups[dom]["total"] += c.get("total") or 0

    field_row = 3
    grand_before = 0
    grand_vat = 0
    grand_total = 0

    for idx, (dom, stats) in enumerate(sorted(field_groups.items()), start=1):
        before = stats["before_vat"]
        vat_amt = stats["vat"]
        total = stats["total"]
        grand_before += before
        grand_vat += vat_amt
        grand_total += total

        ws_field.cell(row=field_row, column=1, value=idx).font = DATA_FONT
        ws_field.cell(row=field_row, column=1).border = BORDER_THIN
        ws_field.cell(row=field_row, column=1).alignment = DATA_ALIGN_CENTER

        ws_field.cell(row=field_row, column=2, value=dom).font = DATA_FONT
        ws_field.cell(row=field_row, column=2).border = BORDER_THIN
        ws_field.cell(row=field_row, column=2).alignment = DATA_ALIGN_LEFT

        ws_field.cell(row=field_row, column=3, value=stats["count"]).font = DATA_FONT
        ws_field.cell(row=field_row, column=3).border = BORDER_THIN
        ws_field.cell(row=field_row, column=3).alignment = DATA_ALIGN_RIGHT

        cell_bv = ws_field.cell(row=field_row, column=4, value=before if before > 0 else 0)
        cell_bv.number_format = '#,##0'
        cell_bv.font = DATA_FONT
        cell_bv.border = BORDER_THIN
        cell_bv.alignment = DATA_ALIGN_RIGHT

        cell_vat = ws_field.cell(row=field_row, column=5, value=vat_amt if vat_amt > 0 else 0)
        cell_vat.number_format = '#,##0'
        cell_vat.font = DATA_FONT
        cell_vat.border = BORDER_THIN
        cell_vat.alignment = DATA_ALIGN_RIGHT

        cell_tot = ws_field.cell(row=field_row, column=6, value=total if total > 0 else 0)
        cell_tot.number_format = '#,##0'
        cell_tot.font = DATA_FONT
        cell_tot.border = BORDER_THIN
        cell_tot.alignment = DATA_ALIGN_RIGHT

        field_row += 1

    # Total row
    if field_groups:
        ws_field.cell(row=field_row, column=1, value="TỔNG CỘNG").font = TOTAL_FONT
        ws_field.cell(row=field_row, column=1).fill = TOTAL_FILL
        ws_field.cell(row=field_row, column=1).border = BORDER_THIN
        ws_field.cell(row=field_row, column=1).alignment = DATA_ALIGN_CENTER

        ws_field.cell(row=field_row, column=2, value="").font = TOTAL_FONT
        ws_field.cell(row=field_row, column=2).fill = TOTAL_FILL
        ws_field.cell(row=field_row, column=2).border = BORDER_THIN

        ws_field.cell(row=field_row, column=3, value=sum(v["count"] for v in field_groups.values())).font = TOTAL_FONT
        ws_field.cell(row=field_row, column=3).fill = TOTAL_FILL
        ws_field.cell(row=field_row, column=3).border = BORDER_THIN
        ws_field.cell(row=field_row, column=3).alignment = DATA_ALIGN_RIGHT

        for col_idx, val in [(4, grand_before), (5, grand_vat), (6, grand_total)]:
            cell = ws_field.cell(row=field_row, column=col_idx, value=val)
            cell.number_format = '#,##0'
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = BORDER_THIN
            cell.alignment = DATA_ALIGN_RIGHT

    _freeze_and_filter(ws_field, row=3)
    _auto_width(ws_field)

    # ── Sheet 4: SAP_HET_HAN ────────────────────────────────────────────
    ws_exp = wb.create_sheet(title="SAP_HET_HAN")

    exp_headers = [
        "STT", "Số HĐ", "Số GCN", "Tên đơn vị", "Lĩnh vực",
        "Phường/Xã sử dụng", "Tỉnh/Thành sử dụng",
        "Ngày hết hạn", "Còn lại (ngày)", "Tiền trước VAT", "VAT", "Tổng tiền", "Người phụ trách",
    ]
    exp_col_widths = [5, 16, 18, 28, 16, 24, 20, 13, 13, 22, 9, 22, 20]

    ws_exp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(exp_headers))
    title_e = ws_exp.cell(row=1, column=1, value=f"HỢP ĐỒNG SẮP HẾT HẠN — {period_label}")
    title_e.font = Font(name="Arial", bold=True, size=13, color="C0392B")
    title_e.alignment = Alignment(horizontal="center", vertical="center")
    ws_exp.row_dimensions[1].height = 22

    _apply_header(ws_exp, 2, exp_headers, exp_col_widths)
    ws_exp.row_dimensions[2].height = 30

    exp_row = 3
    exp_total_before = 0
    exp_total_vat = 0
    exp_total_after = 0

    for idx, contract in enumerate(expiring_contracts, start=1):
        so_tien = contract.get("so_tien_value") or 0
        vat_rate = contract.get("vat_rate") or 0
        vat_amt = contract.get("vat_amount") or 0
        total = contract.get("total") or 0
        days_left = contract.get("days_left", 0)
        expire_date = contract.get("expire_date", "")
        is_warning = days_left <= 30

        exp_total_before += so_tien
        exp_total_vat += vat_amt
        exp_total_after += total

        row_data = [
            idx,
            contract.get("contract_no", "-") or "-",
            contract.get("gcn_no", "-") or "-",
            contract.get("partner", "-") or "-",
            contract.get("field", "-") or "-",
            contract.get("usage_ward") or "-",
            contract.get("usage_province") or "-",
            expire_date,
            days_left if days_left > 0 else "-",
            so_tien if so_tien > 0 else "-",
            f"{vat_rate:.1f}%" if vat_rate else "-",
            total if total > 0 else "-",
            contract.get("nguoi_phu_trach", "-") or "-",
        ]

        _apply_data_row(ws_exp, exp_row, row_data, is_alt=(idx % 2 == 0), is_warning=is_warning)
        format_date_cell(ws_exp.cell(row=exp_row, column=8), expire_date)
        exp_row += 1

    # Total row
    if expiring_contracts:
        merge_end = 5
        ws_exp.merge_cells(start_row=exp_row, start_column=1, end_row=exp_row, end_column=merge_end)
        exp_total_label = ws_exp.cell(row=exp_row, column=1, value="TỔNG CỘNG")
        exp_total_label.font = TOTAL_FONT
        exp_total_label.fill = TOTAL_FILL
        exp_total_label.alignment = DATA_ALIGN_RIGHT
        exp_total_label.border = BORDER_THIN

        for col_idx in range(merge_end + 1, 6):
            ws_exp.cell(row=exp_row, column=col_idx, value="-").border = BORDER_THIN
            ws_exp.cell(row=exp_row, column=col_idx).fill = TOTAL_FILL

        # Ward → col 6
        ws_exp.cell(row=exp_row, column=6, value="-")
        ws_exp.cell(row=exp_row, column=6).font = TOTAL_FONT
        ws_exp.cell(row=exp_row, column=6).fill = TOTAL_FILL
        ws_exp.cell(row=exp_row, column=6).border = BORDER_THIN

        # Province → col 7
        ws_exp.cell(row=exp_row, column=7, value="-")
        ws_exp.cell(row=exp_row, column=7).font = TOTAL_FONT
        ws_exp.cell(row=exp_row, column=7).fill = TOTAL_FILL
        ws_exp.cell(row=exp_row, column=7).border = BORDER_THIN

        # Tiền trước VAT → col 10
        ws_exp.cell(row=exp_row, column=10, value=exp_total_before).number_format = '#,##0'
        ws_exp.cell(row=exp_row, column=10).font = TOTAL_FONT
        ws_exp.cell(row=exp_row, column=10).fill = TOTAL_FILL
        ws_exp.cell(row=exp_row, column=10).alignment = DATA_ALIGN_RIGHT
        ws_exp.cell(row=exp_row, column=10).border = BORDER_THIN

        # VAT % → col 11
        ws_exp.cell(row=exp_row, column=11, value="-")
        ws_exp.cell(row=exp_row, column=11).font = TOTAL_FONT
        ws_exp.cell(row=exp_row, column=11).fill = TOTAL_FILL
        ws_exp.cell(row=exp_row, column=11).alignment = DATA_ALIGN_CENTER
        ws_exp.cell(row=exp_row, column=11).border = BORDER_THIN

        # Tổng tiền → col 12
        ws_exp.cell(row=exp_row, column=12, value=exp_total_after).number_format = '#,##0'
        ws_exp.cell(row=exp_row, column=12).font = TOTAL_FONT
        ws_exp.cell(row=exp_row, column=12).fill = TOTAL_FILL
        ws_exp.cell(row=exp_row, column=12).alignment = DATA_ALIGN_RIGHT
        ws_exp.cell(row=exp_row, column=12).border = BORDER_THIN

        # Người phụ trách → col 13
        ws_exp.cell(row=exp_row, column=13, value="-")
        ws_exp.cell(row=exp_row, column=13).font = TOTAL_FONT
        ws_exp.cell(row=exp_row, column=13).fill = TOTAL_FILL
        ws_exp.cell(row=exp_row, column=13).border = BORDER_THIN

    _freeze_and_filter(ws_exp, row=3)
    _auto_width(ws_exp)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# Revenue Summary Export
# =============================================================================

def export_revenue_summary_xlsx(
    summary_data: dict,
    filters: Optional[dict] = None,
) -> BytesIO:
    """
    Export revenue summary to Excel with multiple sheets.
    
    Args:
        summary_data: Dict with keys:
            - revenue_by_year: list of {year, contract_count, total_revenue}
            - field_breakdown: list of {key, label, count}
            - totals: dict with total contracts, revenue, etc.
        filters: Applied filters for reference
    
    Returns:
        BytesIO buffer with Excel content
    """
    wb = Workbook()
    
    # Sheet 1: Tong_hop
    ws_summary = wb.active
    ws_summary.title = "Tong_hop"
    
    # Headers for summary
    headers_summary = ["Chỉ tiêu", "Giá trị"]
    col_widths_summary = [35, 20]
    
    # Title
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = ws_summary.cell(row=1, column=1, value="BÁO CÁO DOANH THU TỔNG HỢP")
    title_cell.font = Font(name="Arial", bold=True, size=14, color="27AE60")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 25
    
    _apply_header(ws_summary, 2, headers_summary, col_widths_summary)
    
    # Summary data
    totals = summary_data.get("totals", {})
    revenue_data = [
        ("Tổng số hợp đồng", totals.get("total_contracts", 0)),
        ("Hợp đồng đang hoạt động", totals.get("active_count", 0)),
        ("Hợp đồng sắp hết hạn (60 ngày)", totals.get("expiring_60d_count", 0)),
        ("Hợp đồng đã hết hạn", totals.get("expired_count", 0)),
        ("Hợp đồng chờ gia hạn", totals.get("pending_renewal_count", 0)),
        ("Hợp đồng mới", totals.get("new_count", 0)),
    ]
    
    row = 3
    for label, value in revenue_data:
        ws_summary.cell(row=row, column=1, value=label).font = DATA_FONT
        ws_summary.cell(row=row, column=1).border = BORDER_THIN
        ws_summary.cell(row=row, column=1).alignment = DATA_ALIGN_LEFT
        
        if isinstance(value, int) and value > 0:
            ws_summary.cell(row=row, column=2, value=value).number_format = '#,##0'
        else:
            ws_summary.cell(row=row, column=2, value=value)
        ws_summary.cell(row=row, column=2).font = DATA_FONT
        ws_summary.cell(row=row, column=2).border = BORDER_THIN
        ws_summary.cell(row=row, column=2).alignment = DATA_ALIGN_RIGHT
        row += 1
    
    _auto_width(ws_summary)
    
    # Sheet 2: Theo_linh_vuc
    # Note: Phase này chỉ thống kê số lượng HĐ, chưa có revenue breakdown theo lĩnh vực
    ws_field = wb.create_sheet(title="Theo_linh_vuc")
    
    headers_field = ["STT", "Lĩnh vực", "Số lượng HĐ"]
    col_widths_field = [6, 30, 15]
    
    ws_field.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_field = ws_field.cell(row=1, column=1, value="THỐNG KÊ HỢP ĐỒNG THEO LĨNH VỰC")
    title_field.font = Font(name="Arial", bold=True, size=14, color="27AE60")
    title_field.alignment = Alignment(horizontal="center", vertical="center")
    ws_field.row_dimensions[1].height = 25
    
    # Note row
    ws_field.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    note_cell = ws_field.cell(row=2, column=1, value="Ghi chú: Phase này thống kê số lượng HĐ. Doanh thu theo lĩnh vực sẽ bổ sung sau.")
    note_cell.font = Font(name="Arial", italic=True, size=9, color="666666")
    note_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    _apply_header(ws_field, 3, headers_field, col_widths_field)
    ws_field.row_dimensions[3].height = 30
    
    field_data = summary_data.get("field_breakdown", [])
    row = 4  # Start at row 4 (after title, note, header)
    for idx, field in enumerate(field_data, start=1):
        ws_field.cell(row=row, column=1, value=idx).font = DATA_FONT
        ws_field.cell(row=row, column=1).border = BORDER_THIN
        ws_field.cell(row=row, column=1).alignment = DATA_ALIGN_CENTER

        ws_field.cell(row=row, column=2, value=field.get("label", "-")).font = DATA_FONT
        ws_field.cell(row=row, column=2).border = BORDER_THIN
        ws_field.cell(row=row, column=2).alignment = DATA_ALIGN_LEFT

        ws_field.cell(row=row, column=3, value=field.get("count", 0)).font = DATA_FONT
        ws_field.cell(row=row, column=3).border = BORDER_THIN
        ws_field.cell(row=row, column=3).alignment = DATA_ALIGN_RIGHT
        row += 1

    # Total row
    total_field_count = sum(f.get("count", 0) for f in field_data)
    ws_field.cell(row=row, column=1, value="TỔNG CỘNG").font = TOTAL_FONT
    ws_field.cell(row=row, column=1).fill = TOTAL_FILL
    ws_field.cell(row=row, column=1).border = BORDER_THIN
    ws_field.cell(row=row, column=1).alignment = DATA_ALIGN_CENTER

    ws_field.cell(row=row, column=2, value="").font = TOTAL_FONT
    ws_field.cell(row=row, column=2).fill = TOTAL_FILL
    ws_field.cell(row=row, column=2).border = BORDER_THIN

    ws_field.cell(row=row, column=3, value=total_field_count).number_format = '#,##0'
    ws_field.cell(row=row, column=3).font = TOTAL_FONT
    ws_field.cell(row=row, column=3).fill = TOTAL_FILL
    ws_field.cell(row=row, column=3).border = BORDER_THIN
    ws_field.cell(row=row, column=3).alignment = DATA_ALIGN_RIGHT
    
    _freeze_and_filter(ws_field)
    _auto_width(ws_field)
    
    # Sheet 3: Theo_nam
    ws_year = wb.create_sheet(title="Theo_nam")
    
    headers_year = ["Năm", "Số HĐ", "Tổng doanh thu"]
    col_widths_year = [10, 15, 20]
    
    ws_year.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_year = ws_year.cell(row=1, column=1, value="DOANH THU THEO NĂM")
    title_year.font = Font(name="Arial", bold=True, size=14, color="27AE60")
    title_year.alignment = Alignment(horizontal="center", vertical="center")
    ws_year.row_dimensions[1].height = 25
    
    _apply_header(ws_year, 2, headers_year, col_widths_year)
    
    revenue_by_year = summary_data.get("revenue_by_year", [])
    row = 3
    total_contracts = 0
    total_revenue = 0
    for year_data in revenue_by_year:
        year = year_data.get("year", "-")
        count = year_data.get("contract_count", 0)
        revenue = year_data.get("total_revenue") or 0
        
        total_contracts += count
        total_revenue += revenue
        
        ws_year.cell(row=row, column=1, value=year).font = DATA_FONT
        ws_year.cell(row=row, column=1).border = BORDER_THIN
        ws_year.cell(row=row, column=1).alignment = DATA_ALIGN_CENTER
        
        ws_year.cell(row=row, column=2, value=count).font = DATA_FONT
        ws_year.cell(row=row, column=2).border = BORDER_THIN
        ws_year.cell(row=row, column=2).alignment = DATA_ALIGN_RIGHT
        
        ws_year.cell(row=row, column=3, value=revenue).number_format = '#,##0'
        ws_year.cell(row=row, column=3).font = DATA_FONT
        ws_year.cell(row=row, column=3).border = BORDER_THIN
        ws_year.cell(row=row, column=3).alignment = DATA_ALIGN_RIGHT
        row += 1
    
    # Total row
    ws_year.cell(row=row, column=1, value="TỔNG CỘNG").font = TOTAL_FONT
    ws_year.cell(row=row, column=1).fill = TOTAL_FILL
    ws_year.cell(row=row, column=1).border = BORDER_THIN
    ws_year.cell(row=row, column=1).alignment = DATA_ALIGN_CENTER
    
    ws_year.cell(row=row, column=2, value=total_contracts).font = TOTAL_FONT
    ws_year.cell(row=row, column=2).fill = TOTAL_FILL
    ws_year.cell(row=row, column=2).border = BORDER_THIN
    ws_year.cell(row=row, column=2).alignment = DATA_ALIGN_RIGHT
    
    ws_year.cell(row=row, column=3, value=total_revenue).number_format = '#,##0'
    ws_year.cell(row=row, column=3).font = TOTAL_FONT
    ws_year.cell(row=row, column=3).fill = TOTAL_FILL
    ws_year.cell(row=row, column=3).border = BORDER_THIN
    ws_year.cell(row=row, column=3).alignment = DATA_ALIGN_RIGHT
    
    _freeze_and_filter(ws_year)
    _auto_width(ws_year)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# Full Data Export (Tat ca du lieu)
# =============================================================================

def export_full_data_xlsx(
    contracts: list[dict],
    filters: Optional[dict] = None,
    has_music_areas: bool = False,
) -> BytesIO:
    """
    Export full contract data to Excel matching import template columns.

    Columns: Số HĐ, Năm HĐ, Số phụ lục, Tên đơn vị, Địa chỉ đơn vị,
    Điện thoại, Người đại diện, Chức vụ, Mã số thuế, Email,
    Tên biển hiệu, Địa chỉ sử dụng, Địa chỉ pháp lý,
    Phường/Xã pháp lý, Tỉnh pháp lý, Địa chỉ sử dụng đầy đủ,
    Phường/Xã sử dụng, Tỉnh sử dụng, Lĩnh vực, Lĩnh vực hiển thị,
    Ngày lập HĐ, Ngày bắt đầu, Ngày kết thúc, Số tiền (VNĐ),
    % Thuế, Tiền bản quyền trước VAT, Tỷ lệ VAT, Tiền VAT,
    Tiền bản quyền sau VAT, Loại hình karaoke, Tổng số phòng,
    Tổng số box, Người thực hiện, Trạng thái gia hạn,
    Có thể gia hạn, Mẫu hợp đồng

    Plus music_usage_areas columns when has_music_areas=True:
    Khu vực sử dụng âm nhạc, Số phòng/Số ghế/Số lượng,
    Hình thức sử dụng âm nhạc, Ghi chú khu vực

    Args:
        contracts: List of contract dicts with all fields from import template
        filters: Applied filters for reference
        has_music_areas: If True, includes music usage area columns

    Returns:
        BytesIO buffer with Excel content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Du_lieu"

    # Base headers matching import template
    base_headers = [
        "Số hợp đồng", "Năm hợp đồng", "Số phụ lục", "Tên đơn vị",
        "Địa chỉ đơn vị", "Điện thoại", "Người đại diện", "Chức vụ",
        "Mã số thuế", "Email", "Tên biển hiệu", "Địa chỉ sử dụng",
        "Địa chỉ pháp lý", "Phường/Xã pháp lý", "Tỉnh pháp lý",
        "Địa chỉ sử dụng đầy đủ", "Phường/Xã sử dụng", "Tỉnh sử dụng",
        "Lĩnh vực", "Lĩnh vực hiển thị", "Ngày lập hợp đồng",
        "Ngày bắt đầu", "Ngày kết thúc", "Số tiền (VNĐ)", "% Thuế",
        "Tiền bản quyền trước VAT", "Tỷ lệ VAT", "Tiền VAT",
        "Tiền bản quyền sau VAT", "Loại hình karaoke", "Tổng số phòng",
        "Tổng số box", "Người thực hiện", "Trạng thái gia hạn",
        "Có thể gia hạn", "Mẫu hợp đồng"
    ]

    # Music usage area headers
    area_headers = [
        "Vị trí / khu vực sử dụng âm nhạc",
        "Số phòng / số chỗ",
        "Hình thức sử dụng âm nhạc",
        "Ghi chú khu vực",
    ]

    headers = base_headers + area_headers if has_music_areas else base_headers

    # Base field keys matching dict keys from _build_full_data_dict
    base_field_keys = [
        "so_hop_dong", "nam_hop_dong", "so_phu_luc", "ten_don_vi",
        "dia_chi_don_vi", "dien_thoai", "nguoi_dai_dien", "chuc_vu",
        "ma_so_thue", "email", "ten_bien_hieu", "dia_chi_su_dung",
        "dia_chi_phap_ly", "phuong_xa_phap_ly", "tinh_phap_ly",
        "dia_chi_su_dung_day_du", "phuong_xa_su_dung", "tinh_su_dung",
        "linh_vuc", "linh_vuc_hien_thi", "ngay_lap_hop_dong",
        "ngay_bat_dau", "ngay_ket_thuc", "so_tien", "thue_percent",
        "royalty_before_vat", "vat_rate", "vat_amount", "royalty_after_vat",
        "loai_hinh_karaoke", "tong_so_phong", "tong_so_box",
        "nguoi_thuc_hien", "trang_thai_gia_han", "co_the_gia_han", "mau_hop_dong"
    ]

    # Music usage area field keys
    area_field_keys = [
        "area_name",
        "scale_description",
        "music_usage_type",
        "note",
    ]

    field_keys = base_field_keys + area_field_keys if has_music_areas else base_field_keys

    base_col_widths = [15, 10, 12, 30, 35, 15, 20, 12, 15, 25, 20, 35, 35, 15, 15, 35, 15, 15, 15, 15, 15, 15, 15, 15, 15, 12, 12, 12, 12, 15, 12, 10, 20, 15, 12, 15]
    area_col_widths = [30, 22, 28, 25]
    col_widths = base_col_widths + area_col_widths if has_music_areas else base_col_widths
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="BÁO CÁO DỮ LIỆU ĐẦY ĐỦ")
    title_cell.font = Font(name="Arial", bold=True, size=14, color="2D5F8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Filter info row
    if filters:
        filter_text = " | ".join([f"{k}: {v}" for k, v in filters.items() if v])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        filter_cell = ws.cell(row=2, column=1, value=f"Lọc: {filter_text}")
        filter_cell.font = Font(name="Arial", italic=True, size=9, color="666666")
        filter_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    header_row = 3 if filters else 2
    _apply_header(ws, header_row, headers, col_widths)
    ws.row_dimensions[header_row].height = 35
    
    # Data rows
    data_row = header_row + 1
    total_so_tien = 0
    total_royalty_before = 0
    total_vat = 0
    total_royalty_after = 0

    # Date columns indices (1-based): ngay_lap_hop_dong, ngay_bat_dau, ngay_ket_thuc
    # Position in field_keys: base has 36 fields, dates are at positions 21,22,23 (1-based)
    base_date_cols = {21, 22, 23}
    # When has_music_areas=True, area fields add 4 cols, so dates shift to 25,26,27
    date_col_indices = {c + 4 if has_music_areas else c for c in base_date_cols}

    for idx, contract in enumerate(contracts, start=1):
        so_tien = contract.get("so_tien") or 0
        royalty_before = contract.get("royalty_before_vat") or 0
        vat_amt = contract.get("vat_amount") or 0
        royalty_after = contract.get("royalty_after_vat") or 0
        total_so_tien += so_tien
        total_royalty_before += royalty_before
        total_vat += vat_amt
        total_royalty_after += royalty_after

        # Build row data
        row_data = []
        for key in field_keys:
            value = contract.get(key, "")
            row_data.append(value)

        _apply_data_row(ws, data_row, row_data, is_alt=(idx % 2 == 0))

        # Apply date formatting
        for col_idx in date_col_indices:
            field_key = field_keys[col_idx - 1]
            value = contract.get(field_key, "")
            format_date_cell(ws.cell(row=data_row, column=col_idx), value)

        data_row += 1

    # Totals row
    if contracts:
        merge_end = 24 + (4 if has_music_areas else 0)
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=merge_end)
        total_label = ws.cell(row=data_row, column=1, value="TỔNG CỘNG")
        total_label.font = TOTAL_FONT
        total_label.fill = TOTAL_FILL
        total_label.alignment = DATA_ALIGN_RIGHT
        total_label.border = BORDER_THIN

        # Fill empty cells from col merge_end+1 to amount_col-1
        base_col_count = 36 + (4 if has_music_areas else 0)
        amount_col = 25 + (4 if has_music_areas else 0)  # so_tien column

        for col_idx in range(merge_end + 1, amount_col):
            ws.cell(row=data_row, column=col_idx, value="-")
            ws.cell(row=data_row, column=col_idx).font = TOTAL_FONT
            ws.cell(row=data_row, column=col_idx).fill = TOTAL_FILL
            ws.cell(row=data_row, column=col_idx).alignment = DATA_ALIGN_CENTER
            ws.cell(row=data_row, column=col_idx).border = BORDER_THIN

        # so_tien → col 25/29
        ws.cell(row=data_row, column=amount_col, value=total_so_tien).number_format = '#,##0'
        ws.cell(row=data_row, column=amount_col).font = TOTAL_FONT
        ws.cell(row=data_row, column=amount_col).fill = TOTAL_FILL
        ws.cell(row=data_row, column=amount_col).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=amount_col).border = BORDER_THIN

        # thue_percent (col 26/30) → "-"
        ws.cell(row=data_row, column=amount_col + 1, value="-")
        ws.cell(row=data_row, column=amount_col + 1).font = TOTAL_FONT
        ws.cell(row=data_row, column=amount_col + 1).fill = TOTAL_FILL
        ws.cell(row=data_row, column=amount_col + 1).alignment = DATA_ALIGN_CENTER
        ws.cell(row=data_row, column=amount_col + 1).border = BORDER_THIN

        # royalty_before_vat (col 27/31)
        ws.cell(row=data_row, column=amount_col + 2, value=total_royalty_before).number_format = '#,##0'
        ws.cell(row=data_row, column=amount_col + 2).font = TOTAL_FONT
        ws.cell(row=data_row, column=amount_col + 2).fill = TOTAL_FILL
        ws.cell(row=data_row, column=amount_col + 2).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=amount_col + 2).border = BORDER_THIN

        # vat_rate (col 28/32) → "-"
        ws.cell(row=data_row, column=amount_col + 3, value="-")
        ws.cell(row=data_row, column=amount_col + 3).font = TOTAL_FONT
        ws.cell(row=data_row, column=amount_col + 3).fill = TOTAL_FILL
        ws.cell(row=data_row, column=amount_col + 3).alignment = DATA_ALIGN_CENTER
        ws.cell(row=data_row, column=amount_col + 3).border = BORDER_THIN

        # vat_amount (col 29/33)
        ws.cell(row=data_row, column=amount_col + 4, value=total_vat).number_format = '#,##0'
        ws.cell(row=data_row, column=amount_col + 4).font = TOTAL_FONT
        ws.cell(row=data_row, column=amount_col + 4).fill = TOTAL_FILL
        ws.cell(row=data_row, column=amount_col + 4).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=amount_col + 4).border = BORDER_THIN

        # royalty_after_vat (col 30/34)
        ws.cell(row=data_row, column=amount_col + 5, value=total_royalty_after).number_format = '#,##0'
        ws.cell(row=data_row, column=amount_col + 5).font = TOTAL_FONT
        ws.cell(row=data_row, column=amount_col + 5).fill = TOTAL_FILL
        ws.cell(row=data_row, column=amount_col + 5).alignment = DATA_ALIGN_RIGHT
        ws.cell(row=data_row, column=amount_col + 5).border = BORDER_THIN

        # Fill remaining cells
        for col_idx in range(amount_col + 6, base_col_count + 1):
            ws.cell(row=data_row, column=col_idx, value="-")
            ws.cell(row=data_row, column=col_idx).font = TOTAL_FONT
            ws.cell(row=data_row, column=col_idx).fill = TOTAL_FILL
            ws.cell(row=data_row, column=col_idx).alignment = DATA_ALIGN_CENTER
            ws.cell(row=data_row, column=col_idx).border = BORDER_THIN
    
    # Freeze and auto-filter
    _freeze_and_filter(ws, row=header_row + 1)
    _auto_width(ws)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
