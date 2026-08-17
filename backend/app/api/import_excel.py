from __future__ import annotations

import logging
from datetime import date, datetime
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..core.database import get_db
from ..core.security import decode_access_token, security_scheme
from ..models.contracts import ContractRecordRow
from ..services.domain_registry import canonicalize_domain

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/import", tags=["import"])

# Mapping từ header Excel -> tên trường database (cơ bản)
EXCEL_TO_DB_MAPPING: dict[str, str] = {
    "Số hợp đồng": "contract_no",
    "Năm hợp đồng": "contract_year",
    "Số phụ lục": "annex_no",
    "Tên đơn vị": "don_vi_ten",
    "Địa chỉ đơn vị": "don_vi_dia_chi",
    "Điện thoại": "don_vi_dien_thoai",
    "Người đại diện": "don_vi_nguoi_dai_dien",
    "Chức vụ": "don_vi_chuc_vu",
    "Mã số thuế": "don_vi_mst",
    "Email": "don_vi_email",
    "Tên biển hiệu": "ten_bang_hieu",
    "Địa chỉ sử dụng": "dia_chi_su_dung",
    "Địa chỉ pháp lý": "legal_full_address",
    "Phường/Xã pháp lý": "legal_ward",
    "Tỉnh pháp lý": "legal_province",
    "Địa chỉ sử dụng đầy đủ": "usage_full_address",
    "Phường/Xã sử dụng": "usage_ward",
    "Tỉnh sử dụng": "usage_province",
    "Lĩnh vực": "linh_vuc",
    "Ngày lập hợp đồng": "ngay_lap_hop_dong",
    "Ngày bắt đầu": "ngay_bat_dau",
    "Ngày kết thúc": "ngay_ket_thuc",
    "Số tiền (VNĐ)": "so_tien_value",
    "% Thuế": "thue_percent",
    "Tiền bản quyền trước VAT": "royalty_amount_before_vat",
    "Tỷ lệ VAT": "vat_rate",
    "Tiền VAT": "vat_amount",
    "Tiền bản quyền sau VAT": "royalty_amount_after_vat",
    "Loại hình karaoke": "loai_hinh_karaoke",
    "Tổng số phòng": "tong_so_phong",
    "Tổng số box": "tong_so_box",
    "Người thực hiện": "nguoi_thuc_hien_email",
    "Trạng thái gia hạn": "renewal_status",
    "Có thể gia hạn": "is_renewable",
    "Mẫu hợp đồng": "contract_template_code",
}

# Music usage areas columns (Phase 2 - standard field names)
MUSIC_USAGE_AREA_MAPPING: dict[str, str] = {
    "Vị trí / khu vực sử dụng âm nhạc": "area_name",
    "Số phòng / số chỗ": "scale_description",
    "Hình thức sử dụng âm nhạc": "music_usage_type",
    "Ghi chú khu vực": "note",
}

# Legacy Karaoke columns (backward compatible — convert to music_usage_areas)
LEGACY_KARAOKE_MAPPING: dict[str, str] = {
    "Loại hình karaoke": "loai_hinh_karaoke",
    "Tổng số phòng": "tong_so_phong",
    "Tổng số box": "tong_so_box",
}

# Các trường bắt buộc
REQUIRED_FIELDS = ["contract_no", "contract_year"]

# Các trường date cần parse
DATE_FIELDS = ["ngay_lap_hop_dong", "ngay_bat_dau", "ngay_ket_thuc"]

# Các trường số
NUMBER_FIELDS = [
    "contract_year", "so_tien_value", "thue_percent", "royalty_amount_before_vat",
    "vat_rate", "vat_amount", "royalty_amount_after_vat"
]

# Các trường boolean
BOOL_FIELDS = ["is_renewable", "usage_same_as_legal"]

# Những trường legacy Karaoke KHÔNG được ghi mới vào DB (chỉ dùng để convert fallback)
LEGACY_KARAOKE_FIELDS = {"loai_hinh_karaoke", "tong_so_phong", "tong_so_box"}


class ImportResult(BaseModel):
    total_rows: int
    success_count: int
    error_count: int
    errors: list[dict[str, Any]]


def parse_date(value: Any) -> date | None:
    """Parse various date formats."""
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        # Try common formats
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        # Try just year for simple year entries
        if value.isdigit() and len(value) == 4:
            return date(int(value), 1, 1)
    return None


def parse_number(value: Any) -> int | float | None:
    """Parse number from string or numeric value."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        value = value.strip().replace(",", "").replace(" ", "")
        try:
            if "." in value:
                return float(value)
            return int(value)
        except ValueError:
            return None
    return None


def parse_bool(value: Any) -> bool | None:
    """Parse boolean from various representations."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        value = value.strip().lower()
        if value in ["true", "1", "yes", "có", "co", "x", "✓", "✔"]:
            return True
        if value in ["false", "0", "no", "không", "khong"]:
            return False
    return None


def _normalize_usage_method(raw: str) -> str:
    """Normalize usage method to standard values."""
    if not raw:
        return "Phát nhạc nền"
    raw_lower = raw.lower().strip()
    if "karaoke" in raw_lower:
        return "Sử dụng nhạc qua đầu Karaoke"
    if "nền" in raw_lower or "background" in raw_lower:
        return "Phát nhạc nền"
    if "biểu diễn" in raw_lower or "trực tiếp" in raw_lower:
        return "Biểu diễn âm nhạc trực tiếp"
    if "phòng thu" in raw_lower or "thu âm" in raw_lower:
        return "Phòng thu âm"
    return raw.strip()


def _convert_legacy_to_music_areas(
    loai_hinh: str | None,
    tong_so_phong: int | None,
    tong_so_box: int | None,
) -> list[dict]:
    """Convert legacy Karaoke fields to music_usage_areas format."""
    areas = []
    karaoke_type = str(loai_hinh or "").strip().upper()
    if karaoke_type == "BOX" and tong_so_box and tong_so_box > 0:
        areas.append({
            "area_name": "Khu vực sử dụng",
            "scale_description": f"{tong_so_box} box",
            "music_usage_type": "Sử dụng nhạc qua đầu Karaoke (Box)",
            "note": "Converted from legacy fields",
        })
    elif tong_so_phong and tong_so_phong > 0:
        areas.append({
            "area_name": "Khu vực sử dụng",
            "scale_description": f"{tong_so_phong} phòng",
            "music_usage_type": "Sử dụng nhạc qua đầu Karaoke",
            "note": "Converted from legacy fields",
        })
    return areas


def _parse_music_usage_areas_from_row(
    row_data: dict,
    header_to_field: dict[int, str],
    col_values: dict[int, Any],
) -> list[dict]:
    """
    Parse music_usage_areas from a single Excel row.

    If the Excel has music_usage_areas columns, read them.
    If the Excel has legacy Karaoke columns, convert to music_usage_areas.
    Returns a list of area dicts.
    """
    def _get_col(header_to_field: dict[int, str], col_values: dict[int, Any], field_name: str) -> Any:
        for col_idx, fname in header_to_field.items():
            if fname == field_name:
                return col_values.get(col_idx)
        return None

    # Check if this row has music_usage_areas explicit columns
    has_explicit_cols = any(
        fname in header_to_field.values()
        for fname in ["area_name", "scale_description", "music_usage_type", "note"]
    )

    if has_explicit_cols:
        area_name_raw = _get_col(header_to_field, col_values, "area_name")
        scale_raw = _get_col(header_to_field, col_values, "scale_description")
        music_type_raw = _get_col(header_to_field, col_values, "music_usage_type")
        note_raw = _get_col(header_to_field, col_values, "note")

        area_name = str(area_name_raw or "").strip()
        scale_desc = str(scale_raw or "").strip()

        if area_name or scale_desc:
            return [{
                "area_name": area_name,
                "scale_description": scale_desc,
                "music_usage_type": _normalize_usage_method(str(music_type_raw or "")),
                "note": str(note_raw or "").strip(),
            }]

    # Fallback: try legacy Karaoke columns
    loai_hinh = row_data.get("loai_hinh_karaoke")
    tong_so_phong = row_data.get("tong_so_phong")
    tong_so_box = row_data.get("tong_so_box")

    if loai_hinh or tong_so_phong or tong_so_box:
        return _convert_legacy_to_music_areas(loai_hinh, tong_so_phong, tong_so_box)

    return []


@router.post("/contracts", response_model=ImportResult)
async def import_contracts(
    file: UploadFile,
    db: Session = Depends(get_db),
    token: str = Depends(security_scheme),
):
    """
    Import contracts from Excel file.
    
    - Chỉ cần cung cấp các trường muốn import (không bắt buộc đủ)
    - Số hợp đồng và Năm hợp đồng là bắt buộc
    - Nếu hợp đồng đã tồn tại (theo contract_no + contract_year) sẽ update
    """
    user = decode_access_token(token)
    user_role = user.get("role", "")
    
    # Chỉ admin và manager được phép import
    if user_role not in ["admin", "mod"]:
        raise HTTPException(
            status_code=403,
            detail="Chỉ admin và manager mới được phép import dữ liệu"
        )
    
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Chỉ chấp nhận file Excel (.xlsx, .xls)"
        )
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(contents, data_only=True)
        ws = wb.active
        
        # Đọc header từ dòng đầu tiên
        headers: dict[int, str] = {}
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        for idx, cell in enumerate(header_row):
            if cell.value:
                headers[idx] = str(cell.value).strip()
        
        if not headers:
            raise HTTPException(status_code=400, detail="File Excel không có header")
        
        # Map header Excel -> field name
        header_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in EXCEL_TO_DB_MAPPING:
                header_to_field[col_idx] = EXCEL_TO_DB_MAPPING[header_text]
        
        if not header_to_field:
            raise HTTPException(
                status_code=400,
                detail="Không tìm thấy header hợp lệ. Vui lòng sử dụng template chuẩn."
            )
        
        # Xác định các trường bắt buộc có trong file
        present_fields = set(header_to_field.values())
        missing_required = [f for f in REQUIRED_FIELDS if f not in present_fields]
        if missing_required:
            raise HTTPException(
                status_code=400,
                detail=f"Thiếu trường bắt buộc: {', '.join(missing_required)}"
            )
        
        # =================================================================
        # STEP 1: Read ALL data rows into a list first
        # =================================================================
        raw_rows: list[tuple[int, dict[str, Any], dict[int, Any], list[dict]]] = []

        # Also collect music_usage_areas columns from headers
        area_header_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in MUSIC_USAGE_AREA_MAPPING:
                area_header_to_field[col_idx] = MUSIC_USAGE_AREA_MAPPING[header_text]

        # Also track legacy Karaoke columns for fallback conversion
        legacy_header_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in LEGACY_KARAOKE_MAPPING:
                legacy_header_to_field[col_idx] = LEGACY_KARAOKE_MAPPING[header_text]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data: dict[str, Any] = {}
            col_values: dict[int, Any] = {}
            has_data = False

            for col_idx, cell in enumerate(row):
                col_values[col_idx] = cell.value
                if col_idx in header_to_field:
                    field_name = header_to_field[col_idx]
                    value = cell.value

                    if field_name in DATE_FIELDS:
                        value = parse_date(value)
                    elif field_name in NUMBER_FIELDS:
                        value = parse_number(value)
                    elif field_name in BOOL_FIELDS:
                        value = parse_bool(value)

                    if value is not None and value != "":
                        has_data = True

                    row_data[field_name] = value

            if not has_data:
                continue

            # Validate required fields
            row_errors = []
            for req_field in REQUIRED_FIELDS:
                if req_field not in row_data or row_data[req_field] is None or row_data[req_field] == "":
                    row_errors.append(f"Thiếu trường: {req_field}")

            if row_errors:
                raw_rows.append((row_idx, row_data, col_values, row_errors))
                continue

            # Parse music_usage_areas from this row
            # 1. First try music_usage_areas explicit columns
            area_dict: dict[str, Any] = {}
            for col_idx, field_name in area_header_to_field.items():
                raw_val = col_values.get(col_idx)
                if field_name == "music_usage_type":
                    area_dict[field_name] = _normalize_usage_method(str(raw_val) if raw_val else "")
                else:
                    area_dict[field_name] = str(raw_val) if raw_val is not None else ""

            if area_dict.get("area_name") or area_dict.get("scale_description"):
                area_list = [area_dict]
            else:
                # 2. Fallback: convert legacy Karaoke columns
                loai_hinh = row_data.get("loai_hinh_karaoke")
                tong_phong = row_data.get("tong_so_phong")
                tong_box = row_data.get("tong_so_box")
                area_list = _convert_legacy_to_music_areas(loai_hinh, tong_phong, tong_box)

            raw_rows.append((row_idx, row_data, col_values, area_list))

        # =================================================================
        # STEP 2: Group rows by (contract_no, contract_year)
        # =================================================================
        # Map: (contract_no, contract_year) -> {base_data, all_music_areas, row_errors}
        grouped: dict[tuple, dict] = {}
        for row_entry in raw_rows:
            row_idx, row_data, col_values, extra = row_entry
            contract_no = row_data.get("contract_no")
            contract_year = row_data.get("contract_year")
            key = (str(contract_no), contract_year)

            if key not in grouped:
                grouped[key] = {
                    "base_data": row_data,
                    "music_areas": [],
                    "legacy_warnings": [],
                    "row_errors": [],
                    "first_row_idx": row_idx,
                }

            g = grouped[key]

            if isinstance(extra, list) and extra and isinstance(extra[0], dict):
                # This is a music_areas list
                for area in extra:
                    if area:  # Only add non-empty areas
                        g["music_areas"].append(area)
            elif isinstance(extra, list) and extra and isinstance(extra[0], str):
                # This is an error list
                g["row_errors"].extend([(row_idx, e) for e in extra])

        # =================================================================
        # STEP 3: Import each contract
        # =================================================================
        success_count = 0
        error_count = 0
        errors: list[dict[str, Any]] = []

        for (contract_no, contract_year), group_data in grouped.items():
            # Report row-level errors (skip rows without required fields)
            if group_data["row_errors"]:
                for row_idx, err in group_data["row_errors"]:
                    error_count += 1
                    errors.append({"row": row_idx, "errors": [err]})
                continue

            base_data = group_data["base_data"]
            music_areas = group_data["music_areas"]
            legacy_warnings = group_data["legacy_warnings"]

            # Clean up base_data: remove legacy Karaoke fields before saving
            clean_data = {
                k: v for k, v in base_data.items()
                if k not in LEGACY_KARAOKE_FIELDS
            }

            # Add music_usage_areas
            if music_areas:
                import json
                clean_data["music_usage_areas"] = json.dumps(music_areas, ensure_ascii=False)

            # Canonicalize write-boundary inputs so the canonical registry
            # is the only source of truth. Raw labels from Excel must NOT
            # land in business columns without going through this step.
            for _canon_field in ("linh_vuc", "field_code"):
                if _canon_field in clean_data and clean_data[_canon_field]:
                    clean_data[_canon_field] = (
                        canonicalize_domain(clean_data[_canon_field])
                        or clean_data[_canon_field]
                    )

            try:
                existing = db.query(ContractRecordRow).filter(
                    ContractRecordRow.contract_no == contract_no,
                    ContractRecordRow.contract_year == contract_year
                ).first()

                if existing:
                    # Update existing record (only non-None fields)
                    for field, value in clean_data.items():
                        if value is not None and field != "music_usage_areas":
                            setattr(existing, field, value)
                    # Always update music_usage_areas
                    if "music_usage_areas" in clean_data:
                        setattr(existing, "music_usage_areas", clean_data["music_usage_areas"])
                else:
                    # Create new record
                    new_contract = ContractRecordRow(**clean_data)
                    db.add(new_contract)

                db.commit()
                success_count += 1

                if legacy_warnings:
                    errors.append({
                        "row": group_data["first_row_idx"],
                        "errors": [f"Warning: {w}" for w in legacy_warnings]
                    })

            except Exception as e:
                db.rollback()
                error_count += 1
                errors.append({
                    "row": group_data["first_row_idx"],
                    "errors": [f"Lỗi: {str(e)}"]
                })
                logger.error(f"Error importing contract {contract_no}/{contract_year}: {e}")

        return ImportResult(
            total_rows=success_count + error_count,
            success_count=success_count,
            error_count=error_count,
            errors=errors[:100]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý file: {str(e)}")


@router.get("/template")
async def get_template():
    """Download Excel template for import."""
    return {
        "message": "Template available at /templates/contract_import_template.xlsx",
        "headers": list(EXCEL_TO_DB_MAPPING.keys()),
        "required_fields": ["Số hợp đồng", "Năm hợp đồng"],
        "optional_fields": [k for k in EXCEL_TO_DB_MAPPING.keys() 
                           if EXCEL_TO_DB_MAPPING[k] not in REQUIRED_FIELDS]
    }
